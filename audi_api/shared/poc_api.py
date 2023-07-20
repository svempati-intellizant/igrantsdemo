import datetime
import openpyxl
import requests
from pathlib import Path
import os
import logging
from dotenv import load_dotenv
load_dotenv()

absolute_log_path = os.getenv('ABSOLUTE_LOGS_PATH')
logging.basicConfig(filename=absolute_log_path+"/audit.log",
                    format='%(asctime)s %(message)s',
                    filemode='w')
# Creating an object
logger = logging.getLogger()
logger.setLevel(logging.INFO)


def pocApi(file_path):
    logger.info(file_path)
    try:
        file_saved_path = os.getenv('ABSOLUTE_FILE_SAVED_PATH')
        api_url = os.getenv('IGRANT_NEST_API')
        os.makedirs(file_saved_path, exist_ok=True)

        sheet = openpyxl.load_workbook(file_path, data_only=True)

        # Variable declaration
        keys = []
        each_sheet_output = []
        # Iterate first row to get the Headers
        for row in sheet.worksheets[0].iter_rows(max_row=2, min_row=2):
            for cell in row:
                if cell.value != None:
                    keys.append(cell.value)

        # Remove the apostophe in coloumn name
        for i in range(len(keys)):
            keys[i] = keys[i].replace("'", "")
            keys[i] = keys[i].replace('"', '')
        # Iterate other rows to construct the json
        for row in sheet.worksheets[0].iter_rows(min_row=6):
            obj = {}
            for index in range(len(row)):
                if row[index].value != None:
                    if isinstance(row[index].value, str):
                        row[index].value = row[index].value.replace(
                            "'", "")
                        row[index].value = row[index].value.replace(
                            '"', '')
                    if isinstance(row[index].value, datetime.datetime):
                        row[index].value = str(
                            row[index].value)
                    if keys[index] in obj:
                        continue
                    obj[keys[index]] = row[index].value
                else:
                    if keys[index] in obj:
                        continue
                    obj[keys[index]] = ''
            if len(obj)>0:
                each_sheet_output.append(obj)
        output={
            'data':each_sheet_output
        }
        jsonFile = open(f'{file_saved_path}/data.json', "w")
        jsonFile.write(str(output))
        jsonFile.close()

        if len(each_sheet_output)>0:
            logger.info('Can Hit API')
            try:
                logger.info('Hitting API..........')
                with open(f'{file_saved_path}/data.json', 'rb') as f:
                    response = requests.post(os.getenv('IGRANT_NEST_API'), files={'files':f })
                    print(response)
                    if response:
                        return {
                            'error': False,
                            'message': 'Pushed to Database'
                        }
            except Exception as identifier:
                logger.info(f'Failed Hit API..........{identifier}')
                logging.info(
                    f'Failed at processing the file due to {identifier}')
                return{
                    'error': True,
                    'message': "FAILED"
                }

        return {
            'data': each_sheet_output
        }
    except Exception as identifier:
        logging.info(
            f'Failed at processing the file due to {identifier}')
        return{
            'error': True,
            'message': identifier
        }
