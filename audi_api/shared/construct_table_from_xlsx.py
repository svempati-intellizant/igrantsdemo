import datetime
import openpyxl
from pathlib import Path
import os
from dotenv import load_dotenv
load_dotenv()


file_saved_path = os.getenv('ABSOLUTE_FILE_SAVED_PATH')


def sendJsonToIgrantApi(file_path):

    # Get the excel from the path and open the active shee
    sheet = openpyxl.load_workbook(Path(
        file_path), data_only=True)

    # Variable declaration
    tables = []
    keys = []

    # Iterate first row to get the Table Name
    for row in sheet.worksheets[0].iter_rows(max_row=1, min_row=1):
        for cell in row:
            if(cell.value != None):
                table_name = cell.value
                tables.append(table_name)
            else:
                tables.append(table_name)

    # Iterate second row to get the Coloumn Name
    for row in sheet.worksheets[0].iter_rows(max_row=2, min_row=2):
        for cell in row:
            if(cell.value != None):
                keys.append(cell.value)

    current_element = tables[0]
    table_details = []

    # Get the respective start and end index of table
    for i in range(len(tables)):
        if tables[i] == current_element and not i == len(tables)-1:
            continue
        else:
            if len(table_details) == 0:
                table_details.append(
                    {
                        'start_index': 0,
                        'end_index': i-1,
                        'table_name': current_element,
                        'respective_coloumns': []
                    }
                )
            elif i == len(tables)-1:
                table_details.append(
                    {
                        'start_index': table_details[len(table_details)-1]['end_index']+1,
                        'end_index': i-1,
                        'table_name': current_element,
                        'respective_coloumns': []

                    }
                )
            else:
                table_details.append(
                    {
                        'start_index': table_details[len(table_details)-1]['end_index']+1,
                        'end_index': i-1,
                        'table_name': current_element,
                        'respective_coloumns': []

                    }
                )

            current_element = tables[i]

    # Get the Coloumn in Respective Tables
    for i in range(len(table_details)):
        for j in range(len(keys)):
            if j >= table_details[i]['start_index'] and j <= table_details[i]['end_index']:
                table_details[i]['respective_coloumns'].append(keys[j])

    # Remove the apostophe in coloumn name
    for i in range(len(keys)):
        keys[i] = keys[i].replace("'", "")
        keys[i] = keys[i].replace('"', '')

    # Iterate and construct the table with respective rows and coloumns
    for index in range(len(table_details)):
        table_details[index]['coloumns'] = []
        for row in sheet.worksheets[0].iter_rows(min_row=3):
            if row[0].row > 5:
                each_row = {}
                for row_index in range(len(row)):
                    if row_index >= table_details[index]['start_index'] and row_index <= table_details[index]['end_index']:
                        if row[row_index].value != None:

                            if isinstance(row[row_index].value, str):
                                row[row_index].value = row[row_index].value.replace(
                                    "'", "")
                                row[row_index].value = row[row_index].value.replace(
                                    '"', '')
                            if isinstance(row[row_index].value, datetime.datetime):
                                row[row_index].value = str(
                                    row[row_index].value)
                            each_row[keys[row_index]] = row[row_index].value
                        else:
                            each_row[keys[row_index]] = ''
                table_details[index]['coloumns'].append(each_row)
    return table_details
